import os
import json
import pytest


import json
import os
import pytest
from commonlib import config, time_utils, sign_utils
from commonlib.log_utils import log
from commonlib.api_lib.base_api import BaseApi
from commonlib.api_lib.AES import get_encrypt_info,get_decrypt_info
from commonlib.api_lib.validator import schema_validator,digit_number_validator,idc_verify
from commonlib.api_lib.AES_new import *


from defines.belt.identity_service_business import IdentitySwaggerBusiness
from defines.belt.ocr_service_business import OcrSwaggerBusiness
from defines.belt.face_service_business import FaceSwaggerBusiness
from commonlib.sign_utils import encode_jwt_token


import sys
print(sys.path)

import openpyxl

class Accuracy:
    def __init__(self,config):
        #host
        self.host=config["host"]
        #ak
        self.ak=config["ak"]
        #sk
        self.sk=config["sk"]
        #真值数据文件
        self.ground_truth_file=config["ground_truth_file"]
        #测试数据文件，包括正负样本
        self.test_image_file=config["test_image_file"]
        #真值数据
        self.ground_truth_data=None
        #测试数据
        self.test_image_data=None
        #结果数据
        self.detect_result={}
        #初始化检测子
        self.OcrApi=None
        #测试结果存储路径
        self.result_dir=f"/Users/weishuting/accuracy_result"
        #测试结果文件名称
        self.result_name="ocr_idcard.xlsx"
        #最终测试结果路径
        self.excel_save_path=None
        #测试结果excel的sheet_name
        self.SHEET_NAME=f"测试结果"
        #文件头
        self.excel_header=["测试数据","检出name","实际name","TP（完全匹配）","FP(存在错误)","FN(没检出)","检出number","实际number","TP（完全匹配）","FP(存在错误)","FN(没检出)","备注"]
        #汇总信息头
        self.excel_sum_header=["检出项","样本总数","TP（完全匹配）","FP(存在错误)","FN(没检出)","召回（TP/总数）","准确率(TP/(FP+FP))"]
        #写入行数据索引，初始值为0
        self.row_index=0
    def init_excle(self):
        """初始化表头"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        #测试结果文件带时间戳
        if not os.path.exists(self.result_dir):
            os.makedirs(self.result_dir)
        import datetime
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        result_name=str(nowTime)+"_"+self.result_name
        self.excel_save_path=os.path.join(self.result_dir,result_name)        
        sheet.title =self.SHEET_NAME
        self.row_index=self.row_index+1
        for i, val in enumerate(self.excel_header):
            sheet.cell(row=self.row_index, column=i+1, value=val)
        workbook.save(self.excel_save_path)

    def write_row_data(self,row_data):
        """按行写入数据"""
        self.row_index=self.row_index+1
        workbook = openpyxl.load_workbook(self.excel_save_path)
        sheet = workbook[self.SHEET_NAME]
        for i, val in enumerate(row_data):
            sheet.cell(row=self.row_index, column=i+1, value=val)
        workbook.save(self.excel_save_path)

    def calculte(self):
        """计算精度指标，输入真值和检测结果，得到精度数据"""
        #初始化表头
        self.init_excle()
        #汇总结果初始化
        name_SUM_TP=0
        name_SUM_FP=0
        name_SUM_FN=0 
        number_SUM_TP=0
        number_SUM_FP=0
        number_SUM_FN=0 
        image_SUM=0

        #写入详细结果
        for image_name,detect_result in self.detect_result.items():
            excel_each_row_data=[]
            if image_name in self.ground_truth_data.keys():
                image_SUM=image_SUM+1
                excel_each_row_data.append(image_name)
                #TP,FP,FN,写入逻辑
                #name字段
                TP=0
                FP=0
                FN=0
                excel_each_row_data.append(detect_result["name"])
                excel_each_row_data.append(self.ground_truth_data[image_name]["name"])
                if detect_result["name"]==self.ground_truth_data[image_name]["name"]:
                    TP=1
                    name_SUM_TP=name_SUM_TP+TP
                elif detect_result["name"]=="":
                    FN=1
                    name_SUM_FN=name_SUM_FN+FN                 
                elif detect_result["name"]!=self.ground_truth_data[image_name]["name"]:
                    FP=1
                    name_SUM_FP=name_SUM_FP+FP
                    
                excel_each_row_data.append(TP)
                excel_each_row_data.append(FP)
                excel_each_row_data.append(FN)

                #number字段
                TP=0
                FP=0
                FN=0
                excel_each_row_data.append(detect_result["number"])
                excel_each_row_data.append(self.ground_truth_data[image_name]["number"])
                if detect_result["number"]==self.ground_truth_data[image_name]["number"]:
                    TP=1
                    number_SUM_TP=number_SUM_TP+TP
                elif detect_result["number"]=="":
                    FN=1
                    number_SUM_FN=number_SUM_FN+FN                   
                elif detect_result["number"]!=self.ground_truth_data[image_name]["number"]:
                    FP=1
                    number_SUM_FP=number_SUM_FP+FP

                excel_each_row_data.append(TP)
                excel_each_row_data.append(FP)
                excel_each_row_data.append(FN)
                #写入一个测试图片的测试结果信息
                self.write_row_data(excel_each_row_data)
            else:
                print("该图片：%s"%image_name+"无对应的标注数据！！！")        
        
        #写入汇总信息头
        self.write_row_data(self.excel_sum_header)
        # 写入汇总信息
        #name
        excel_sum_row_data=[]          
        excel_sum_row_data.append("name")
        excel_sum_row_data.append(image_SUM)
        excel_sum_row_data.append(name_SUM_TP)
        excel_sum_row_data.append(name_SUM_FP)
        excel_sum_row_data.append(name_SUM_FN)
        recall=float(name_SUM_TP/image_SUM)
        accuracy=float(name_SUM_TP/(name_SUM_TP+name_SUM_FP))
        excel_sum_row_data.append(recall)
        excel_sum_row_data.append(accuracy)
        self.write_row_data(excel_sum_row_data)
        #number
        excel_sum_row_data=[]          
        excel_sum_row_data.append("number")
        excel_sum_row_data.append(image_SUM)
        excel_sum_row_data.append(number_SUM_TP)
        excel_sum_row_data.append(number_SUM_FP)
        excel_sum_row_data.append(number_SUM_FN)
        recall=float(number_SUM_TP/image_SUM)
        accuracy=float(number_SUM_TP/(number_SUM_TP+number_SUM_FP))
        excel_sum_row_data.append(recall)
        excel_sum_row_data.append(accuracy)
        self.write_row_data(excel_sum_row_data)
        return 
    def init_detecter(self):
        """初始化单接口检测子"""
        aksk_token=encode_jwt_token(self.ak,self.sk)
        self.OcrApi=OcrSwaggerBusiness(self.host, token=aksk_token)    
    def detect_run(self):
        """检测函数，输入数据到接口，获取检测结果""" 
        video_paths_t=self.test_image_data["image_path_t"]
        for image_path_t in video_paths_t:
            base64_image=str(BaseApi.imageToBase64(image_path_t))
            side = 'FRONT'
            detect_quality = False
            encrypt_info={
                "algorithm": "ENCRPT_ALGORITHM_NONE",
                "version": 0,
                "encrypted_fields": [
                "string"
                ],
                "data": "string"
            } 
            if "elsezu" in  image_path_t:
                wst=0
                wst=wst+1            
            resp = self.OcrApi.OCRService_OCRIDCardPostApi(image=base64_image, side=side, detect_quality=detect_quality, encrypt_info=encrypt_info)
            dirname,full_file_name=os.path.split(image_path_t)
            #构造检出
            value={}
            #如果无name和number字段检出，赋值为“”
            if "error" in resp.resp_json:
                print(f"接口调用发生错误！！！")
                value={
                "name": "error",
                "number":"error"
                }

            if "idcard" not in resp.resp_json:
                value={
                "name": "",
                "number":""
                }
                print(f"检出无idcard字段！！！")
            else:
                if "name" not in resp.resp_json["idcard"].keys():
                    value["name"]=""
                    print(f"检出无name字段！！！")
                else:
                    value["name"]=resp.resp_json["idcard"]["name"]
                if "number" not in resp.resp_json["idcard"].keys():
                    value["number"]=""
                    print(f"检出无number字段！！！")
                else:
                    value["number"]=resp.resp_json["idcard"]["number"]       
            self.detect_result[full_file_name]=value
            
    def get_ground_truth_data_from_file(self):
        """获取真值数据，输入真值文件，得到真值数据"""
        with open(self.ground_truth_file,"r") as f:
            self.ground_truth_data=json.load(f)
        return

    def get_test_imgae_data_from_file(self):
        """获取测试数据，输入测试数据文件，得到测试数据"""
        with open(self.test_image_file,"r") as f:
            self.test_image_data=json.load(f)
        return    

    def run(self):
        """入口函数"""
        #获取真值数据
        self.get_ground_truth_data_from_file()
        #获取测试数据
        self.get_test_imgae_data_from_file()
        #初始化检测子
        self.init_detecter()
        #获取检测结果
        self.detect_run()
        #计算结果
        self.calculte()


if __name__=="__main__":
    #输入1--标注文件
    ground_truth_file="/Users/weishuting/scp/viper/data/身份证/2022-12-28-送标文件-done/ground_truth.json"
    #输入2--素材
    test_image_file="/Users/weishuting/scp/viper/data/身份证/2022-12-28-送标文件-done/test_image.json"
    #输入3--host
    # host="https://alb-dcc-1-cn-shanghai-1.sensebelt.com"
    # host="https://ids.test.sensebelt.net"
    # #实例化
    # config={
    #     "ground_truth_file":ground_truth_file,
    #     "test_image_file":test_image_file,
    #     "host":host,
    #     "ak":"2OXzMGrVjTWMtbwYlM6L3avxswt",
    #     "sk":"D7Tba3EpuweR69R2UBF38yG4BtTj9Hut"
    # }

    host="https://ids.sensebelt.com"
    #实例化
    config={
        "ground_truth_file":ground_truth_file,
        "test_image_file":test_image_file,
        "host":host,
        "ak" :"2PdikkghZm9pA4ITFY3YrvhQpya",
        "sk" : "ltMiPwCtz0d5e1XakOlnlLOHO0kX8gil"
    }   
    # host="https://ids.staging.sensebelt.com"
    # #实例化
    # config={
    #     "ground_truth_file":ground_truth_file,
    #     "test_image_file":test_image_file,
    #     "host":host,
    #     "ak" : "2PdcT4neMhutHx8deahGvpPodgd",
    #     "sk" :"yfNzHagXqUD20RNiHXpIBqrgemgemYDw"
    # } 


    runner=Accuracy(config=config)
    #开始
    ret=runner.run()

    

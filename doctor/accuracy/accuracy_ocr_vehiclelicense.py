import os
import json
import pytest


import json
import os
import pytest
from commonlib import config, time_utils, sign_utils
from commonlib.log_utils import log
from commonlib.api_lib.base_api import BaseApi
from commonlib.api_lib.AES import get_encrypt_info,get_decrypt_info
from commonlib.api_lib.validator import schema_validator,digit_number_validator,idc_verify
from commonlib.api_lib.AES_new import *


from defines.belt.identity_service_business import IdentitySwaggerBusiness
from defines.belt.ocr_service_business import OcrSwaggerBusiness
from defines.belt.face_service_business import FaceSwaggerBusiness
from commonlib.sign_utils import encode_jwt_token


import sys
print(sys.path)

import openpyxl

class Accuracy:
    def __init__(self,config):
        #host
        self.host=config["host"]
        #ak
        self.ak=config["ak"]
        #sk
        self.sk=config["sk"]
        #真值数据文件
        self.ground_truth_file=config["ground_truth_file"]
        #测试数据文件，包括正负样本
        self.test_image_file=config["test_image_file"]
        #真值数据
        self.ground_truth_data=None
        #测试数据
        self.test_image_data=None
        #结果数据
        self.detect_result={}
        #初始化检测子
        self.OcrApi=None
        #测试结果存储路径
        self.result_dir=f"/Users/weishuting/accuracy_result"
        #测试结果文件名称
        self.result_name="ocr_vehiclelicense.xlsx"
        #最终测试结果路径
        self.excel_save_path=None
        #测试结果excel的sheet_name
        self.SHEET_NAME=f"测试结果"
        #文件头
        self.excel_header=["测试数据","检出owner","实际owner","owner TP（完全匹配","owner FP(存在错误)","owner FN(没检出)","检出plate_no","实际plate_no","plate_no TP（完全匹配","plate_no FP(存在错误)","plate_no FN(没检出)","备注"]
        #汇总信息头
        self.excel_sum_header=["检出项","样本总数","TP（完全匹配）","FP(存在错误)","FN(没检出)","召回（TP/总数）","准确率(TP/(FP+TP))"]
        #写入行数据索引，初始值为0
        self.row_index=0
    def init_excle(self):
        """初始化表头"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        #测试结果文件带时间戳
        if not os.path.exists(self.result_dir):
            os.makedirs(self.result_dir)
        import datetime
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        result_name=str(nowTime)+"_"+self.result_name
        self.excel_save_path=os.path.join(self.result_dir,result_name)        
        sheet.title =self.SHEET_NAME
        self.row_index=self.row_index+1
        for i, val in enumerate(self.excel_header):
            sheet.cell(row=self.row_index, column=i+1, value=val)
        workbook.save(self.excel_save_path)

    def write_row_data(self,row_data):
        """按行写入数据"""
        self.row_index=self.row_index+1
        workbook = openpyxl.load_workbook(self.excel_save_path)
        sheet = workbook[self.SHEET_NAME]
        for i, val in enumerate(row_data):
            sheet.cell(row=self.row_index, column=i+1, value=val)
        workbook.save(self.excel_save_path)

    def calculte(self):
        """计算精度指标，输入真值和检测结果，得到精度数据"""
        #初始化表头
        self.init_excle()
        #汇总结果初始化
        name_SUM_TP=0
        name_SUM_FP=0
        name_SUM_FN=0 

        name_SUM_TP_1=0
        name_SUM_FP_1=0
        name_SUM_FN_1=0 

        image_SUM=0

        #写入详细结果
        for image_name,detect_result in self.detect_result.items():
            excel_each_row_data=[]
            if image_name in self.ground_truth_data.keys():
                image_SUM=image_SUM+1
                excel_each_row_data.append(image_name)
                #TP,FP,FN,写入逻辑
                #owner字段
                TP=0
                FP=0
                FN=0
                excel_each_row_data.append(detect_result["owner"])
                excel_each_row_data.append(self.ground_truth_data[image_name]["owner"])
                if detect_result["owner"]==self.ground_truth_data[image_name]["owner"]:
                    TP=1
                    name_SUM_TP=name_SUM_TP+TP
                elif detect_result["owner"]=="":
                    FN=1
                    name_SUM_FN=name_SUM_FN+FN                 
                elif detect_result["owner"]!=self.ground_truth_data[image_name]["owner"]:
                    FP=1
                    name_SUM_FP=name_SUM_FP+FP
                    
                excel_each_row_data.append(TP)
                excel_each_row_data.append(FP)
                excel_each_row_data.append(FN)

                #plate_no字段
                TP_1=0
                FP_1=0
                FN_1=0
                excel_each_row_data.append(detect_result["plate_no"])
                excel_each_row_data.append(self.ground_truth_data[image_name]["plate_no"])
                if detect_result["plate_no"]==self.ground_truth_data[image_name]["plate_no"]:
                    TP_1=1
                    name_SUM_TP_1=name_SUM_TP_1+TP_1
                elif detect_result["plate_no"]=="":
                    FN_1=1
                    name_SUM_FN_1=name_SUM_FN_1+FN_1                 
                elif detect_result["plate_no"]!=self.ground_truth_data[image_name]["plate_no"]:
                    FP_1=1
                    name_SUM_FP_1=name_SUM_FP_1+FP_1
                    
                excel_each_row_data.append(TP_1)
                excel_each_row_data.append(FP_1)
                excel_each_row_data.append(FN_1)
                #写入一个测试图片的测试结果信息
                self.write_row_data(excel_each_row_data)
            else:
                print("该图片：%s"%image_name+"无对应的标注数据！！！")        
        
        #写入汇总信息头
        self.write_row_data(self.excel_sum_header)
        # 写入汇总信息
        #owner
        excel_sum_row_data=[]          
        excel_sum_row_data.append("owner")
        excel_sum_row_data.append(image_SUM)
        excel_sum_row_data.append(name_SUM_TP)
        excel_sum_row_data.append(name_SUM_FP)
        excel_sum_row_data.append(name_SUM_FN)
        recall=float(name_SUM_TP/image_SUM)
        accuracy=float(name_SUM_TP/(name_SUM_TP+name_SUM_FP))
        excel_sum_row_data.append(recall)
        excel_sum_row_data.append(accuracy)
        self.write_row_data(excel_sum_row_data)
        #plate_no
        excel_sum_row_data=[]          
        excel_sum_row_data.append("plate_no")
        excel_sum_row_data.append(image_SUM)
        excel_sum_row_data.append(name_SUM_TP_1)
        excel_sum_row_data.append(name_SUM_FP_1)
        excel_sum_row_data.append(name_SUM_FN_1)
        recall=float(name_SUM_TP_1/image_SUM)
        accuracy=float(name_SUM_TP_1/(name_SUM_TP_1+name_SUM_FP_1))
        excel_sum_row_data.append(recall)
        excel_sum_row_data.append(accuracy)   
        self.write_row_data(excel_sum_row_data)
        return 
    def init_detecter(self):
        """初始化单接口检测子"""
        aksk_token=encode_jwt_token(self.ak,self.sk)
        self.OcrApi=OcrSwaggerBusiness(self.host, token=aksk_token)    
    def detect_run(self):
        """检测函数，输入数据到接口，获取检测结果""" 
        video_paths_t=self.test_image_data["image_path_t"]
        for image_path_t in video_paths_t:
            encrypt_info={
                "algorithm": "ENCRPT_ALGORITHM_NONE",
                "version": 0,
                "encrypted_fields": [
                "string"
                ],
                "data": "string"
            }
            base64_image=str(BaseApi.imageToBase64(image_path_t))
            resp = self.OcrApi.OCRService_OCRVehicleLicensePostApi(image=base64_image, encrypt_info=encrypt_info)
            dirname,full_file_name=os.path.split(image_path_t)
            #构造检出
            value={}
            #如果无name和number字段检出，赋值为“”
            if "error" in resp.resp_json:
                print(f"接口调用发生错误！！！")
                value={
                "plate_no": "error",
                "owner":"error"
                }

            if "first_page" not in resp.resp_json:
                value={
                "plate_no": "",
                "owner":""
                }
                print(f"检出无first_page字段！！！")
            else:
                if "plate_no" not in resp.resp_json["first_page"].keys():
                    value["plate_no"]=""
                    print(f"检出无plate_no字段！！！")
                else:
                    value["plate_no"]=resp.resp_json["first_page"]["plate_no"]     
                if "owner" not in resp.resp_json["first_page"].keys():
                    value["owner"]=""
                    print(f"检出无owner字段！！！")
                else:
                    value["owner"]=resp.resp_json["first_page"]["owner"]     



            self.detect_result[full_file_name]=value
            
    def get_ground_truth_data_from_file(self):
        """获取真值数据，输入真值文件，得到真值数据"""
        with open(self.ground_truth_file,"r") as f:
            self.ground_truth_data=json.load(f)
        return

    def get_test_imgae_data_from_file(self):
        """获取测试数据，输入测试数据文件，得到测试数据"""
        with open(self.test_image_file,"r") as f:
            self.test_image_data=json.load(f)
        return    

    def run(self):
        """入口函数"""
        #获取真值数据
        self.get_ground_truth_data_from_file()
        #获取测试数据
        self.get_test_imgae_data_from_file()
        #初始化检测子
        self.init_detecter()
        #获取检测结果
        self.detect_run()
        #计算结果
        self.calculte()


if __name__=="__main__":
    #输入1--标注文件
    ground_truth_file="/Users/weishuting/code/belt-beta-new/doctor/resource/images/ids_image/ocr/vehicle_license/ground_truth.json"
    #输入2--素材
    test_image_file="/Users/weishuting/code/belt-beta-new/doctor/resource/images/ids_image/ocr/vehicle_license/test_image.json"
    #输入3--host
    # host="https://alb-dcc-1-cn-shanghai-1.sensebelt.com"
    # host="https://crd-test.sensetime.com/ids-wrapper"
    # #实例化
    # config={
    #     "ground_truth_file":ground_truth_file,
    #     "test_image_file":test_image_file,
    #     "host":host,
    #     "ak":"2JA6G2FFP5qoLIaZC1GDDmdgI5E",
    #     "sk":"Zd3APGzZeEUjnVnKle5YwZstS2R1Epp6"
    # }
    # runner=Accuracy(config=config)
    # #开始
    # ret=runner.run()

    

    host="https://ids.test.sensebelt.net"
    #实例化
    config={
        "ground_truth_file":ground_truth_file,
        "test_image_file":test_image_file,
        "host":host,
        "ak":"2OXzMGrVjTWMtbwYlM6L3avxswt",
        "sk":"D7Tba3EpuweR69R2UBF38yG4BtTj9Hut"
    }
    runner=Accuracy(config=config)
    #开始
    ret=runner.run()

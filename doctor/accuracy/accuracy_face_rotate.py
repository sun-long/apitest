import json
import os
import pytest
from commonlib import config, time_utils, sign_utils
from commonlib.log_utils import log
from commonlib.api_lib.base_api import BaseApi
from commonlib.api_lib.AES import get_encrypt_info,get_decrypt_info
from commonlib.api_lib.validator import schema_validator,digit_number_validator,idc_verify
from commonlib.api_lib.AES_new import *


from defines.belt.identity_service_business import IdentitySwaggerBusiness
from defines.belt.ocr_service_business import OcrSwaggerBusiness
from defines.belt.face_service_business import FaceSwaggerBusiness
from commonlib.sign_utils import encode_jwt_token


import sys
print(sys.path)

import openpyxl

class Accuracy:
    def __init__(self,config):
        #host
        self.host=config["host"]
        #ak
        self.ak=config["ak"]
        #sk
        self.sk=config["sk"]
        #真值数据文件
        self.ground_truth_file=config["ground_truth_file"]
        #测试数据文件，包括正负样本
        self.test_image_file=config["test_image_file"]
        #真值数据
        self.ground_truth_data=None
        #测试数据
        self.test_image_data=None
        #结果数据
        self.detect_result={}
        #初始化检测子
        self.OcrApi=None
        #测试结果存储路径
        self.result_dir=f"/Users/weishuting/accuracy_result"
        #测试结果文件名称
        self.result_name="face_rotate.xlsx"
        #最终测试结果路径
        self.excel_save_path=None
        #测试结果excel的sheet_name
        self.SHEET_NAME=f"测试结果"
        #文件头
        self.excel_header=["测试数据","是否入库成功（0/1）","入库结果返回","备注"]
        #汇总信息头
        self.excel_sum_header=["检出项","样本总数","成功入库率（入库成功数/总数）"]
        #写入行数据索引，初始值为0
        self.row_index=0
    def init_excle(self):
        """初始化表头"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        #测试结果文件带时间戳
        if not os.path.exists(self.result_dir):
            os.makedirs(self.result_dir)
        import datetime
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        result_name=str(nowTime)+"_"+self.result_name
        self.excel_save_path=os.path.join(self.result_dir,result_name)        
        sheet.title =self.SHEET_NAME
        self.row_index=self.row_index+1
        for i, val in enumerate(self.excel_header):
            sheet.cell(row=self.row_index, column=i+1, value=val)
        workbook.save(self.excel_save_path)

    def write_row_data(self,row_data):
        """按行写入数据"""
        self.row_index=self.row_index+1
        workbook = openpyxl.load_workbook(self.excel_save_path)
        sheet = workbook[self.SHEET_NAME]
        for i, val in enumerate(row_data):
            sheet.cell(row=self.row_index, column=i+1, value=val)
        workbook.save(self.excel_save_path)

    def calculte(self):
        """计算精度指标，输入真值和检测结果，得到精度数据"""
        #初始化表头
        self.init_excle()
        #汇总结果初始化
        name_SUM_TP=0
        image_SUM=0
        #写入详细结果
        for image_name,detect_result in self.detect_result.items():
            excel_each_row_data=[]
            image_SUM=image_SUM+1
            excel_each_row_data.append(image_name)
            
            excel_each_row_data.append(detect_result["add_db"])
            excel_each_row_data.append(json.dumps(detect_result["response"]))
            if detect_result["add_db"]==1:
                name_SUM_TP=name_SUM_TP+1
                #写入一个测试图片的测试结果信息
                self.write_row_data(excel_each_row_data)
            else:
                print("该图片：%s"%image_name+"无对应的标注数据！！！")        
        
        #写入汇总信息头
        self.write_row_data(self.excel_sum_header)
        # 写入汇总信息
        #name
        excel_sum_row_data=[]          
        excel_sum_row_data.append("face_rotate")
        excel_sum_row_data.append(image_SUM)
        acc=float(name_SUM_TP/image_SUM)
        excel_sum_row_data.append(acc)
        self.write_row_data(excel_sum_row_data)
        return 
    def init_detecter(self):
        """初始化单接口检测子"""
        aksk_token=encode_jwt_token(self.ak,self.sk)
        self.FaceApi=FaceSwaggerBusiness(host, token=aksk_token)
        self.db_id=self.FaceApi.newDB()
    def tear_down(self):
        """清除环境"""
        #删除库
        self.FaceApi.FaceService_DeletePersonDBPostApi(id=self.db_id)
    def detect_run(self):
        """检测函数，输入数据到接口，获取检测结果""" 
        video_paths_t=self.test_image_data["image_path_t"]
        index=0
        for image_path_t in video_paths_t:
            auto_rotate=True
            index=index+1
            person_id = "a"+str(index)
            image=str(BaseApi.imageToBase64(image_path_t))
            images=[image]
            resp = self.FaceApi.FaceService_AddPersonPostApi(db_id=self.db_id, person_id=person_id, images=images, extra_info=None, auto_rotate=auto_rotate, min_quality_level=None, tag_ids=None, encrypt_info=None)
            value={
                "add_db":"",
                "response":""
            }
            if resp.status_code == 200:
                value["add_db"]=1
            else:
                value["add_db"]=0
            value["response"]=resp.resp_json   
            dirname,full_file_name=os.path.split(image_path_t)
            self.detect_result[full_file_name]=value
            
            
    def get_ground_truth_data_from_file(self):
        """获取真值数据，输入真值文件，得到真值数据"""
        if self.ground_truth_file:
            with open(self.ground_truth_file,"r") as f:
                self.ground_truth_data=json.load(f)
            return

    def get_test_imgae_data_from_file(self):
        """获取测试数据，输入测试数据文件，得到测试数据"""
        if  self.test_image_file:
            with open(self.test_image_file,"r") as f:
                self.test_image_data=json.load(f)
            return    

    def run(self):
        """入口函数"""
        #获取真值数据
        self.get_ground_truth_data_from_file()
        #获取测试数据
        self.get_test_imgae_data_from_file()
        #初始化检测子
        self.init_detecter()
        #获取检测结果
        self.detect_run()
        #计算结果
        self.calculte()
        self.tear_down()


if __name__=="__main__":
    #输入1--标注文件
    ground_truth_file=None
    #输入2--素材
    test_image_file="/Users/weishuting/scp/viper/data/人脸_旋转/test_image.json"
    #输入3--host
    host="https://ids.test.sensebelt.net"
    #实例化
    config={
        "ground_truth_file":ground_truth_file,
        "test_image_file":test_image_file,
        "host":host,
        "ak":"2OXzMGrVjTWMtbwYlM6L3avxswt",
        "sk":"D7Tba3EpuweR69R2UBF38yG4BtTj9Hut"
    }
    runner=Accuracy(config=config)
    #开始
    ret=runner.run()

    
